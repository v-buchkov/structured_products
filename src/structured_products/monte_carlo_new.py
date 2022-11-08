# Ctrl+Shift+Fn+F10

import numpy as np
import scipy
import scipy.linalg
import openpyxl
from tqdm import tqdm
from scipy.stats.stats import pearsonr
import random
import os
from data.default_table import default_table
from data.spx import spx

# MC conditions
# drift is assumed at the most recent dividend yield. Can take (!) analyst recommendation level or AM prediction
# input [True] to take historical mean
mean_dict={'pd':[0.04, 0.06], 'copper':[0.03], 'esg':[True], 'mrna':[0.0], 'nvax':[0.0],
           'robotics':[0.054], 'sxpp':[True], 'wm': [0.016, 0.007, 0.015], 'ls': [0.05, 0.05],
           'rbi': [0.05, 0.05], 'pimco': [True], 'm&g': [True],
           'ltq1': [0.025, 0.01, 0.008, 0.013, 0.009, 0.02], 'gazprom': [0.1],
           'lux_clothes': [0.013, 0.02, 0.006], 'lux_brands': [0.013, 0.007, 0.02], 'lux': [0.015, 0.02, 0.00],
           'space': [0.004, 0.000], 'networks': [0.027, 0.00, 0.00],
           'lux_multi': [0.028, 0.032, 0.023, 0.02, 0.036, 0.077],
           'games': [0.04094, 0.04244, 0.04175], 'cybersec': [0.04178, 0.0377, 0.03767], 'nuclear_cpn': [0.0492],
           'nuclear_kg': [0.0492, 0.0466, 0.0496, 0.0502], 'clear_enr': [0.05, 0.0859, 0.0742],
           'fast_food': [0.048, 0.0525, 0.0577], 'lux_indiv': [0.028, 0.032, 0.0579, 0.02, 0.036, 0.077]}
paths=10000

# SP conditions, cpn p.a.
underlying='fast_food'
years=1
# autocall in years (e.g., 0.25 for quarterly)
autocall=0
coupon=5
barrier=65

worst=1
autocall_start=1
autocall_level=100
cap=0
gearing=1
bonus=0
memory=0
digital=0
digital_t=0
cpn_barrier=barrier
# set True for most recent value - 'XX.XX.XXXX' otherwise
fixing=True
# set True for most recent value - 'XX.XX.XXXX' otherwise
simulation_start=True
issuer_rating='A-'
tail_event=55
compare=False

# ------------------------------------------------------------------------------------------
t=int(252*years)

if years>=1:
    default_p=default_table[issuer_rating][years-1]/100
else:
    default_p=0

n_stocks=len(os.listdir(path='../data/{}'.format(underlying)))

# read data => generate list [[stock1_value1, stock1_value2, ...], [stock2_value1, stock2_value2, ...], ...]
stocks=[]
start_values = []
st_with_dates=[]
skip=7
hp={}
ws=(openpyxl.load_workbook('../data/{}/{}.xlsx'.format(underlying, 1))).active.values
list=[item for item in ws]
with_dates={}
# record the start values
start_values.append(float(list[skip+1][1]))
for i in range(skip+2, len(list)):
    if list[i][0] is None:
        break
    else:
        # list_.append([item[0].strftime("%d.%m.%Y"), float(item[1])])
        try:
            with_dates[list[i-1][0].strftime("%d.%m.%Y")]=[float(list[i-1][1]), i-1]
            hp[list[i-1][0].strftime("%d.%m.%Y")]=[(float(list[i-1][1])/float(list[i][1])-1)]
        except:
            with_dates[list[i-1][0].strftime("%d.%m.%Y")]=[0, i-1]
            hp[list[i - 1][0].strftime("%d.%m.%Y")]=[0]
st_with_dates.append(with_dates)

for k in range(2, n_stocks+1):
    with_dates={}
    ws=(openpyxl.load_workbook('../data/{}/{}.xlsx'.format(underlying, k))).active.values
    list=[item for item in ws]
    # record the start values
    start_values.append(float(list[skip+1][1]))
    for i in range(skip+2, len(list)):
        if list[i][0] is None:
            break
        else:
            # list_.append([item[0].strftime("%d.%m.%Y"), float(item[1])])
            try:
                with_dates[list[i-1][0].strftime("%d.%m.%Y")]=[float(list[i-1][1]), i-1]
                hp[list[i-1][0].strftime("%d.%m.%Y")].append((float(list[i-1][1])/float(list[i][1])-1))
            except:
                with_dates[list[i-1][0].strftime("%d.%m.%Y")]=[0, i-1]
    st_with_dates.append(with_dates)

for i in range(n_stocks):
    stocks.append([])

for item in hp.values():
    if len(item)==n_stocks:
        for i in range(len(item)):
            stocks[i].append(item[i])

# take only until fixing
if simulation_start!=True:
    stocks = [stocks[i][st_with_dates[i][simulation_start][1]-skip-1:] for i in range(len(stocks))]
    start_values=[item[simulation_start][0] for item in st_with_dates]

if fixing!=True:
    fixing_values=[item[fixing][0] for item in st_with_dates]
else:
    fixing_values=start_values

# reduce dataset to the shortest lifetime of all stocks
stocks=[item[:min([len(it) for it in stocks])] for item in stocks]

for item in stocks:
    item.reverse()

# calculate historical mean return
if any(mean_dict[underlying])==True:
    mean_returns=[]
    for stock in stocks:
        mean_returns.append(np.mean(stock)*252)
else:
    mean_returns=mean_dict[underlying]

# calculate var-covar matrix
var_covar=[]
for i in range(n_stocks):
    line=[]
    for j in range(n_stocks):
        line.append(252*sum([(stocks[i][k]-np.mean(stocks[i]))*(stocks[j][k]-np.mean(stocks[j])) for k in range(len(stocks[i]))])/(len(stocks[i])-1))
    var_covar.append(line)

# var_covar=np.cov(stocks)
#
# [print(cell) for cell in var_covar]
#
# var_covar=[[18.813/100, 2.425/100, 2.579/100], [2.425/100, 3.597/100, 3.315/100], [2.579/100, 3.315/100, 4.731/100]]
#
# [print(cell) for cell in var_covar]

# correlation matrix to check
corr=[]
for i in range(len(var_covar)):
    line=[]
    for j in range(len(var_covar)):
        line.append(round(var_covar[j][i]/((var_covar[j][j])**(0.5)*(var_covar[i][i])**(0.5)), 2))
    corr.append(line)

[print(cell) for cell in corr]

wb = openpyxl.Workbook()
ws = wb.active

print('---')
# output=[[stock1_path1, stock2_path1, ..., stock6_path1], [stock1_path2, stock2_path2, ..., stock6_path2], ...]
m=[[] for i in range(n_stocks)]
r=[[] for i in range(n_stocks)]
while True:
    output=[]
    time=np.linspace(0, t/252, t)
    print('Generating paths...')
    cholesky = np.linalg.cholesky(np.array(var_covar))
    for k in tqdm(range(paths)):
        path = [[item] for item in start_values]
        for i in range(1, len(time)):
            dif=np.matmul(cholesky, np.random.normal(0, 1, size=n_stocks))*np.sqrt(time[i]-time[i-1])
            # print(np.exp(dif))
            drift = [(mean_returns[j] - 0.5 * var_covar[j][j]) * (time[i] - time[i - 1]) for j in range(n_stocks)]
            # print(np.exp(drift))
            factor=[np.exp(drift[k]+dif[k]) for k in range(n_stocks)]
            for j in range(n_stocks):
                path[j].append(path[j][i-1]*factor[j])
                r[j].append(factor[j]-1)
            # print(dif)
            # for j in range(n_stocks):
            #     drift = (mean_returns[j] - 0.5 * var_covar[j][j])*(time[i]-time[i-1])
            #     path[j].append(path[j][i-1]*np.exp(drift + dif[j]*((time[i]-time[i-1])**0.5)))
        for j in range(n_stocks):
            m[j].append(path[j][-1]/path[j][0]-1)
        output.append(path)
    print('---')
    print([np.mean(item) for item in m])
    print([np.mean(item) for item in r])

    for_graph=[]
    for path in output:
        returns = []
        for stock in path:
            returns.append((stock[-1] / stock[0] - 1) * 100)
        returns.sort()
        restr = [item for item in returns[:worst]]
        for_graph.append(round(np.mean(restr), 2))

    for i in range(-100, 400):
        k = 0
        for r in for_graph:
            if r > i and r <= i + 1:
                k += 1
        ws['A{}'.format(i + 101)] = round(((2 * i + 1) / 2) / 100, 4)
        ws['B{}'.format(i + 101)] = k

    x=random.randrange(0, paths-6)
    for i in range(0, len(output[0][0])):
        ws['D{}'.format(i + 1)] = round(output[x][0][i], 2)
        ws['E{}'.format(i + 1)] = round(output[x+1][0][i], 2)
        ws['F{}'.format(i + 1)] = round(output[x+2][0][i], 2)
        ws['G{}'.format(i + 1)] = round(output[x+3][0][i], 2)
        ws['H{}'.format(i + 1)] = round(output[x+4][0][i], 2)
        ws['I{}'.format(i + 1)] = round(output[x+5][0][i], 2)
        ws['J{}'.format(i + 1)] = round(output[x+6][0][i], 2)

    wb.save('MC_paths.xlsx')

    # calculate corr new to check
    #
    # corr_new=[]
    # for i in range(n_stocks):
    #     line = []
    #     for j in range(n_stocks):
    #         c=0
    #         for path in output:
    #             c += pearsonr(path[i], path[j])[0]
    #         line.append(round(c/len(output), 2))
    #     corr_new.append(line)
    #
    # [print(cell) for cell in corr]
    # print('---')
    # [print(cell) for cell in corr_new]

    # calculate var_covar new to check
    #
    # var_covar_new=[]
    # for i in range(n_stocks):
    #     line = []
    #     for j in range(n_stocks):
    #         c=0
    #         for path in output:
    #             # c += np.cov(path[i], path[j])
    #             print(np.cov(path[i], path[j]))
    #         # line.append(round(c/len(output), 2))
    #     var_covar_new.append(line)
    #
    # [print(cell) for cell in var_covar]
    # print('---')
    # [print(cell) for cell in var_covar_new]

    # start calculating returns (autocall dependent)
    note_returns=[]
    c=[0, 0, 0, 0]
    lo=[]
    dur = 0
    mean=0
    tail=0
    print('Calculating returns...')
    if memory!=0:
        for i in range(len(output)):
            if n_stocks == 1:
                mean += (output[i][0][-1] / output[i][0][0] - 1) * 100 / paths
            c_ = False
            cpn_sum=0
            for y in range(1, int(t / 252 / autocall)):
                returns = []
                for j in range(len(output[i])):
                    returns.append((output[i][j][int(y*autocall*252)-1] / fixing_values[j] - 1) * 100)
                returns.sort()
                restr = [item for item in returns[:worst]]
                if all(it>=autocall_level-100 for it in restr):
                    note_returns.append(memory * y * autocall)
                    c_ = True
                    dur += (memory * y * autocall + 100) * y * autocall
                    c[3] += 1 / paths
                    break
                elif all(it>=cpn_barrier-100 for it in restr):
                    cpn_sum+=int(memory/(t / 252 / autocall))
            if not c_:
                returns = []
                for j in range(len(output[i])):
                    returns.append((output[i][j][-1] / fixing_values[j] - 1) * 100)
                returns.sort()
                restr = [item for item in returns[:worst]]
                if all(it<barrier-100 for it in restr):
                    note_returns.append(cpn_sum + returns[0])
                    dur += (cpn_sum + returns[0] + 100) * int(t / 252)
                    lo.append(cpn_sum + returns[0])
                    c[0] += 1 / paths
                    if all(it < tail_event - cpn_sum - 100 for it in restr):
                        tail+=1/paths
                else:
                    note_returns.append(cpn_sum)
                    dur += (cpn_sum) * int(t / 252)
                    c[1] += 1 / paths
    elif digital!=0:
        for i in range(len(output)):
            if n_stocks == 1:
                mean += (output[i][0][-1] / output[i][0][0] - 1) * 100 / paths
            cpn_sum=0
            for y in range(1, int(t / 252 / digital_t)+1):
                returns = []
                for j in range(len(output[i])):
                    returns.append((output[i][j][int(y * digital_t * 252)-1] / fixing_values[j] - 1) * 100)
                returns.sort()
                restr = [item for item in returns[:worst]]
                if all(it>=cpn_barrier-100 for it in restr):
                    cpn_sum+=digital*digital_t
            returns = []
            for j in range(len(output[i])):
                returns.append((output[i][j][-1] / fixing_values[j] - 1) * 100)
            returns.sort()
            restr = [item for item in returns[:worst]]
            if all(it<barrier-100 for it in restr):
                note_returns.append(cpn_sum + returns[0])
                c[0] += 1 / paths
                if all(it < tail_event - cpn_sum - 100 for it in restr):
                    tail+=1/paths
            else:
                note_returns.append(cpn_sum)
                c[1] += 1 / paths
    elif autocall!=0:
        for i in range(len(output)):
            if n_stocks == 1:
                mean += (output[i][0][-1] / output[i][0][0] - 1) * 100 / paths
            c_ = False
            for y in range(autocall_start, int(t / 252 / autocall)):
                returns = []
                for j in range(len(output[i])):
                    returns.append((output[i][j][int(y * autocall * 252)-1] / fixing_values[j] - 1) * 100)
                returns.sort()
                restr = [item for item in returns[:worst]]
                if all(it>=autocall_level-100 for it in restr):
                    note_returns.append(coupon * y * autocall)
                    c_ = True
                    dur += (coupon * y * autocall + 100) * y * autocall
                    c[3] += 1 / paths
                    break
            if not c_:
                returns = []
                for j in range(len(output[i])):
                    returns.append((output[i][j][-1] / fixing_values[j] - 1) * 100)
                # print('{}-{}'.format(output[i][0][-1], fixing_values[0]))
                returns.sort()
                restr = [item for item in returns[:worst]]
                if all(it<barrier-100 for it in restr):
                    note_returns.append(coupon * int(t / 252) + returns[0])
                    dur += (coupon * int(t / 252) + returns[0] + 100) * int(t / 252)
                    lo.append(coupon * int(t / 252) + returns[0])
                    c[0] += 1 / paths
                    if all(it < tail_event - coupon * int(t / 252) - 100 for it in restr):
                        tail+=1/paths
                else:
                    note_returns.append(coupon * int(t / 252))
                    dur += (coupon * int(t / 252) + 100) * int(t / 252)
                    c[1] += 1 / paths
    elif bonus!=0:
        for i in range(len(output)):
            returns = []
            if n_stocks == 1:
                mean += (output[i][0][-1] / output[i][0][0] - 1) * 100 / paths
            for j in range(len(output[i])):
                returns.append((output[i][j][-1] / fixing_values[j] - 1) * 100)
            returns.sort()
            restr = [item for item in returns[:worst]]
            if all(it<barrier-100 for it in restr):
                note_returns.append(returns[0]+coupon * int(t / 252))
                lo.append(returns[0]+coupon * int(t / 252))
                c[0]+=1/paths
                if all(it < tail_event - 100 for it in restr):
                    tail += 1 / paths
            elif all(it<bonus for it in restr):
                note_returns.append(bonus+coupon * int(t / 252))
                c[1]+=1/paths
            elif all(it>cap for it in restr):
                note_returns.append(cap*gearing+coupon * int(t / 252))
                c[1]+=1/paths
            else:
                note_returns.append(returns[0]*gearing+coupon * int(t / 252))
                c[2]+=1/paths
    elif cap!=0:
        for i in range(len(output)):
            if n_stocks == 1:
                mean += (output[i][0][-1] / output[i][0][0] - 1) * 100 / paths
            returns = []
            for j in range(len(output[i])):
                returns.append((output[i][j][-1] / fixing_values[j] - 1) * 100)
            returns.sort()
            restr = [item for item in returns[:worst]]
            if all(it<barrier-100 for it in restr):
                note_returns.append(returns[0]+coupon*int(t/252))
                lo.append(returns[0]+coupon*int(t/252))
                c[0]+=1/paths
                if all(it < tail_event - coupon * int(t / 252) - 100 for it in restr):
                    tail += 1 / paths
            elif all(it<0 for it in restr):
                note_returns.append(coupon * int(t / 252))
                c[1]+=1/paths
            elif all(it>cap for it in restr):
                note_returns.append(cap*gearing+coupon * int(t / 252))
                c[2]+=1/paths
            else:
                note_returns.append(returns[0]*gearing+coupon * int(t / 252))
                c[2]+=1/paths
    else:
        for i in range(len(output)):
            if n_stocks == 1:
                mean += (output[i][0][-1] / output[i][0][0] - 1) * 100 / paths
            returns = []
            for j in range(len(output[i])):
                returns.append((output[i][j][-1] / fixing_values[j] - 1) * 100)
            returns.sort()
            restr = [item for item in returns[:worst]]
            if all(it<barrier-100 for it in restr):
                note_returns.append(returns[0]+coupon*int(t/252))
                lo.append(returns[0]+coupon*int(t/252))
                c[0]+=1/paths
                if all(it < tail_event - coupon * int(t / 252) - 100 for it in restr):
                    tail += 1 / paths
            elif all(it<0 for it in restr):
                note_returns.append(coupon * int(t / 252))
                c[1]+=1/paths
            else:
                note_returns.append(returns[0]*gearing+coupon * int(t / 252))
                c[2]+=1/paths
    for note in range(int(paths*default_p)):
        note_returns[random.randrange(0, paths)]=-50
    # if compare:
    #     spx_res = spx(paths, years, autocall, coupon, barrier, worst, autocall_level, cap, gearing, bonus, memory, digital,
    #               cpn_barrier, fixing, simulation_start, issuer_rating, tail_event)
    #     print('{}% mean return vs S&P {}% ({}% vs {}% annualized)'.format(round(np.mean(note_returns), 2), round(
    #         100 * (1 + np.mean(note_returns) / 100) ** (1 / years) - 100, 2),
    #                                                                       spx_res[0], spx_res[1]))
    # else:
    print('---')
    print('Results:')
    print('{}% mean return ({}% annualized)'.format(round(np.mean(note_returns), 2), round(
            100 * (1 + np.mean(note_returns) / 100) ** (1 / years) - 100, 2)))
    if barrier>0 and len(lo)>0:
        print('{}% recovery'.format(round(100+np.mean(lo)), 2))
    if autocall!=0:
        dur = dur / (sum(note_returns) + 100*len(note_returns))
        if dur!=0:
            if dur>1:
                print('Expected duration {} years'.format(round(dur, 2)))
            else:
                print(
                    'Expected duration {} months'.format(round(12 * dur, 0)))
    print('---')
    # print('{}% loss vs S&P {}% loss'.format(round(100*c[0], 2), spx_res[2]))
    print('{}% loss'.format(round(100 * c[0], 2)))
    print('{}% just coupon'.format(round(100*c[1], 2)))
    print('{}% participation'.format(round(100*c[2], 2)))
    print('{}% autocall'.format(round(100 * c[3], 2)))
    print('---')
    print('{}% probability of {}% tail'.format(round(100 * tail, 2), tail_event))
    print('---')
    if n_stocks==1:
        print('{}% stock mean return'.format(round(mean, 2)))
        print('---')
    x=input('Run again? (y/n): ')
    if x=='n':
        break